#自動化_關聯測試_取API指定值
import requests
import json
import pandas
from pandas import DataFrame
import csv
import numpy
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,colors,Border,Side,Alignment
import openpyxl

#函式 FindExtrctKey()，尋找指定Key在哪個欄位中第二層
def FindExtrctKey(find_dict, extract_value):
    for key in find_dict.keys():
        if type(find_dict[key]) == list and find_dict[key] != []: 
            if type(find_dict[key][0]) == dict:
                if extract_value in find_dict[key][0]:
                    return key

#函式 Extract()，取指定值
def Extract(response_dict, extract_value):
    value_list = []
    #如果data是array
    if type(response_dict["data"]) == list:
        #如果指定值在第一層
        if extract_value in response_dict["data"][0]:
            for data_dict in response_dict["data"]:
                value_list.append(data_dict[extract_value])
        #如果指定值在第二層
        else:
            key = FindExtrctKey(response_dict["data"][0], extract_value)
            if key != None:
                for data_dict in response_dict["data"]:
                    for key_dict in data_dict[key]:
                        value_list.append(key_dict[extract_value])
    #如果data非array
    else:
        key = FindExtrctKey(response_dict["data"], extract_value)
        if key != None:
            for key_dict in response_dict["data"][key]:
                value_list.append(key_dict[extract_value])
    return value_list

#函式 ExtractValue()，使用Id或者Ids取指定值，判斷Error情形與回傳value_list取值結果
def ExtractValue(API_url, post_data, extract_value, two_dim_to_one = True, remove_duplicate_flag = True):
    value_list = []
    response = call_API(API_url, post_data)
    try:
        response_dict = response.json()
    except:
        print(API_url)
        print(post_data)
        print(response)
    try:
        value_list = Extract(response_dict, extract_value)
        #找尋無結果，改用Id+"s"查詢
        if value_list == []:
            value_list = Extract(response_dict, extract_value+"s")
            if two_dim_to_one != False:
                temp_list = []
                for list_data in value_list:
                    temp_list = temp_list + list_data
                value_list = temp_list
        #去除重複值
        if remove_duplicate_flag != False:
            temp_list = []
            for i in value_list:
                if not i in temp_list:
                    temp_list.append(i)
            value_list = temp_list
    #data為空值
    except IndexError:
        value_list = []
        #print("Error：No Data Extract")
    #其他問題
    except:
        print("Error：Extract Something Wrong at API:" + API_url)
    #return value_list
    return value_list[:20]

def call_API(API_url, post_data):
    url = domain_url + API_url
    response = requests.post(url, headers = header, json = post_data)
    return response

def next_and_compare(value_list, post_data, extract_key, API_url, work_sheet):
    for i in range(len(value_list)):
        work_sheet.cell(row = i+2, column = 1, value = str(i+1))
        fail_count = 0
        fail_message = ""
        #依據value_list更新post_data
        post_data[extract_key] = value_list[i]
        work_sheet.cell(row = i+2, column = 2, value = json.dumps(post_data, indent = 4))
        #API_Next Request
        next_response = call_API(API_url, post_data)
        try:
            next_dict = next_response.json()
        except:
            print(API_url)
            print(post_data)
            print(next_response)
            continue
        #ExtractValue()將指定Key存在next_list，供以下比對        
        #two_dim_to_one = False 因為API_Next要比對各Request結果，所以不將二維list轉為一維
        #remove_duplicate_flag = False 因為API_Next要比對所有結果，所以不將重複值移除
        next_list = ExtractValue(API_url, post_data, extract_key, False, False)
        #API基本資訊
        #http response status
        work_sheet.cell(row = i+2, column = 3, value = next_response.status_code)
        if next_response.status_code != 200:
            fail_message = fail_message + "Expected http response status is 200 but we got [" + str(next_response.status_code) + "]\n"
        #response body code
        work_sheet.cell(row = i+2, column = 4, value = next_dict["code"])
        if next_dict["code"] != 0:
            fail_message = fail_message + "Expected response body code is 0 but we got [" + str(next_dict["code"]) + "]\n"
        #response body message
        work_sheet.cell(row = i+2, column = 5, value = next_dict["message"])
        if next_dict["message"] != "success":
            fail_message = fail_message + "Expected response body message is [success] but we got [" + next_dict["message"] + "]\n"
        #比對next_list資料
        #如果next_list結果為空集合，代表測試為Block(無法驗證)，以下不用判斷直接跳過此for迴圈
        if next_list == []:
            #print("Block")
            work_sheet.cell(row = i+2, column = 6, value = "Block")
            #print(extract_key + " = " + str(value_list[i]) + ", Result: Block")
            #print("No Data Can Compare")
            work_sheet.cell(row = i+2, column = 7, value = "No Data Can Compare")
            work_sheet.cell(row = i+2, column = 8, value = "[]")
            continue
        #如果API_Next為Ids，以二維list判斷各list是否包含指定值
        elif type(next_list[0]) == list:
            for next_response in next_list:
                if value_list[i] not in next_response:
                    fail_count = fail_count+1
        #如果API_Next為Id，以list判斷各值是否等於指定值
        else:
            for next_response in next_list:            
                if value_list[i] != next_response:
                    fail_count = fail_count+1
        #如果fail_count > 0則該測試案例失敗
        if fail_count:
            #fail_message = fail_message + extract_key + " = " + str(value_list[i]) + ", Compare Result: Fail\n"
            fail_message = fail_message + "Error: " + str(fail_count) + "/" + str(len(next_list)) + "\n"
        #如果fail_message有錯誤訊息，判斷測試結果為Fail反之為Pass，記錄測試數據
        if fail_message:
            #print("Fail")
            work_sheet.cell(row = i+2, column = 6, value = "Fail")
            #print(fail_message[:-1])
            work_sheet.cell(row = i+2, column = 7, value = fail_message[:-1])
            work_sheet.cell(row = i+2, column = 8, value = str(next_list))
        else:
            #print("Pass")
            work_sheet.cell(row = i+2, column = 6, value = "Pass")
            #print(extract_key + " = " + str(value_list[i]) + ", Compare Result: Pass")
            #print("Total: " + str(len(next_list)))
            work_sheet.cell(row = i+2, column = 7, value = "Total: " + str(len(next_list)))
            work_sheet.cell(row = i+2, column = 8, value = str(next_list))

def ExtractMidValue(value_list, post_data, extract_key, extract_next_key, API_url):
    mid_list = []
    for i in range(len(value_list)):
        #依據value_list更新post_data
        post_data[extract_key] = value_list[i]
        #API_Next Request
        next_response = call_API(API_url, post_data)
        try:
            next_dict = next_response.json()
        except:
            print(API_url)
            print(post_data)
            print(next_response)
            continue
        #ExtractValue()將指定Key存在next_list
        next_list = ExtractValue(API_url, post_data, extract_next_key)
        mid_list = mid_list + next_list
    temp_list = []
    for i in mid_list:
        if not i in temp_list:
            temp_list.append(i)
    mid_list = temp_list
    #print(mid_list)
    #return mid_list
    return mid_list[:20]

def loadExcelAPI(file_name, col, sheet):
    #excel路徑,usecols:第幾列,keep_default_na:NAN取消顯示
    df = pandas.read_excel(file_name, usecols = [col], keep_default_na = False ,sheet_name = sheet)
    #空白轉NAN,inplace: True 在原資料修改, False新增新資料在修改,原資料不改
    spaceToNan = df.replace('', numpy.nan, inplace=True)
    #刪除NAN值的欄位
    df.dropna(subset=spaceToNan, inplace=True)
    #轉成array
    dfToArray = numpy.array(df)
    #轉成List,此為多個List
    ArrayToList = dfToArray.tolist()
    #多個List合併成1個List
    oneListApiPre= [x for j in ArrayToList for x in j]
    #刪除List奇數位
    k = 0
    for i in range(len(oneListApiPre)):
        if i % 2 == 1:
            del oneListApiPre[k]
            k += 1            
    return oneListApiPre

def loadExcelValue(file_name, col, sheet):
    #excel路徑,usecols:第幾列,keep_default_na:NAN取消顯示
    df = pandas.read_excel(file_name, usecols = [col], keep_default_na = False ,sheet_name = sheet)
    #空白轉NAN,inplace: True 在原資料修改, False新增新資料在修改,原資料不改
    spaceToNan = df.replace('', numpy.nan, inplace=True)
    #刪除NAN值的欄位
    df.dropna(subset=spaceToNan, inplace=True)
    #轉成array
    dfToArray = numpy.array(df)
    #轉成List,此為多個List
    ArrayToList = dfToArray.tolist()
    #多個List合併成1個List
    oneListValue= [x for j in ArrayToList for x in j]    
    return oneListValue

def loadExcelAPIAll(file_name, Link_num):
    API = []
    for i in range(Link_num):
        API.append(loadExcelAPI(file_name, (i*2)+1, Link_num-2))
    temp_list = []
    for i in range(len(API[0])):
        for j in range(Link_num):
            temp_list.append(API[j][i])
    return temp_list

def loadExcelValueAll(file_name, Link_num):
    value = []
    for i in range(Link_num-1):
        value.append(loadExcelValue(file_name, (i*2)+2, Link_num-2))
    temp_list = []
    for i in range(len(value[0])):
        for j in range(Link_num-1):
            temp_list.append(value[j][i])
    return temp_list

def postDataFormat(work_sheet, testcase_num):
    title = (["No.","Post Data","Status","Code","Message","Result","Note","Linked Value"])
    for i in range(len(title)):
        work_sheet.cell(row = 1, column = i+1, value = title[i])
    work_sheet.column_dimensions['B'].width = 50.0
    work_sheet.column_dimensions['G'].width = 30.0
    work_sheet.column_dimensions['H'].width = 70.0
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    fill = openpyxl.styles.PatternFill("solid", fgColor= "ADD8E6")
    align = Alignment(horizontal='center',vertical='center')
    align2 = Alignment(vertical='center')
    border = Border(bottom=bottom)
    border2 = Border(left=left)

    for row in range(1, testcase_num):
        for i in 'ABCDEFGH':
            work_sheet[i + str(row)].border = border
            work_sheet[i + str(row)].alignment = align2

    for i , row in enumerate(work_sheet.columns):
        cell1 = row[0]
        cell1.fill = fill
        cell1.alignment = align

    for leftRow in range(1,testcase_num):
        for i in 'I':
            work_sheet[i + str(leftRow)].border = border2
            
    for i in range(testcase_num):
        work_sheet.cell(row = i+2, column = 2).alignment = Alignment(wrap_text=True)

#參數
#domain_url設定，Beta環境
domain_url = "https://infoapi.raccoontv.com"
#header設定，application/json
header = {
    "Content-Type": "application/json",
    "Cookie": "__cfduid=d060a840dc0e723def59398968ee53a1b1607311355"
}
wb = load_workbook(r"D:\關聯test\測試案例_LoL_關聯_210113.xlsx")
#excel讀入
#二項關聯
#API_url_all = ["/api/v1/esport/lol/league/list", "/api/v1/esport/lol/team/list", "/api/v1/esport/lol/league/list", "/api/v1/esport/lol/team/awards", "/api/v1/esport/lol/league/list", "/api/v1/esport/lol/league_hero/stats"]
#extract_value_all = ["leagueId", "leagueId", "leagueId"]
API_url_all = loadExcelAPIAll(r"D:\關聯test\測試案例_LoL_關聯_210113.xlsx", 2)
extract_value_all = loadExcelValueAll(r"D:\關聯test\測試案例_LoL_關聯_210113.xlsx", 2)
for i in range(len(extract_value_all)):
    if i < 9:
        ws = wb.create_sheet('02_0' + str(i+1))
    else:
        ws = wb.create_sheet('02_' + str(i+1))
    #post_data預設，token
    post_data = {
        "token": "5TEV2u7T8nFH3Ri78iC0SbubWK80bl9y"
    }
    value_list = ExtractValue(API_url_all[i*2], post_data, extract_value_all[i])
    next_and_compare(value_list, post_data, extract_value_all[i], API_url_all[(i*2)+1], ws)
    postDataFormat(ws, len(value_list)+2)
#三項關聯
API_url_all = loadExcelAPIAll(r"D:\關聯test\測試案例_LoL_關聯_210113.xlsx", 3)
extract_value_all = loadExcelValueAll(r"D:\關聯test\測試案例_LoL_關聯_210113.xlsx", 3)
for i in range(int(len(extract_value_all)/2)):
    if i < 9:
        ws = wb.create_sheet('03_0' + str(i+1))
    else:
        ws = wb.create_sheet('03_' + str(i+1))
    #post_data預設，token
    post_data = {
        "token": "5TEV2u7T8nFH3Ri78iC0SbubWK80bl9y"
    }
    pre_value_list = ExtractValue(API_url_all[i*3], post_data, extract_value_all[i*2])
    #print(pre_value_list)
    mid_value_list = ExtractMidValue(pre_value_list, post_data, extract_value_all[(i*2)], extract_value_all[(i*2)+1], API_url_all[(i*3)+1])
    post_data = {
        "token": "5TEV2u7T8nFH3Ri78iC0SbubWK80bl9y"
    }
    next_and_compare(mid_value_list, post_data, extract_value_all[(i*2)+1], API_url_all[(i*3)+2], ws)
    postDataFormat(ws, len(mid_value_list)+2)
#excel寫出
wb.save(r"D:\關聯test\測試案例_LoL_關聯_210113_test_1.xlsx")